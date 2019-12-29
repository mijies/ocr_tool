import openpyxl
import os

# This is a hyper API for openpyxl

def new_excel_handle(file_path):
    if '/' in file_path:
        file_path = file_path.replace('/', os.sep)

    if os.path.exists(file_path):
        raise IOError(file_path + ': already exists')

    openpyxl.Workbook().save(file_path)
    return ExcelHandle(file_path, writable=True)


# decorater for writable methods
def is_writable(func):
    def wrapper(*args):
        if not args[0].writable:
            raise NotImplementedError('Writable methods not allowed in readonly mode')
        func(*args)
        args[0].book.save(args[0].path)
    return wrapper


class ExcelHandle:
    def __init__(self, file_path, writable=False):
        self.path = file_path
        self.book = load_excel(file_path)
        self.writable = writable


    # return all sheetnames
    def get_sheetnames(self):
        return self.book.sheetnames


    # return active sheetname
    def get_active_sheetname(self):
        return self.book.active.title


    # return a list of sheet names
    def find_sheets(self, rows, columns, func):
        def fn(sheetname):
            address = self.find_cell(sheetname, rows, columns, func)
            return address != (None, None)
        return list(filter(fn, self.book.sheetnames))


    # return a tuple of (row, column)
    def find_cell(self, sheetname, rows, columns, func):
        sheet = self.book[sheetname]
        for row in rows:
            for col in columns:
                if func(sheet.cell(row=row, column=col).value):
                    return (row, col)
        return (None, None)


    # return a list of target cell valus by generator
    def iterate_row_values(self, sheetname, rows, col, func, stop_func=lambda x: True):
        sheet = self.book[sheetname]
        while stop_func(col):
            values = [sheet.cell(row=row, column=col).value for row in rows]
            if func(values):
                break
            yield values
            col += 1

    
    ### Methods only when writable is True

    # add a new sheet
    @is_writable
    def add_sheet(self, name, index=None):
        kwargs = {'title': name}
        if index is not None:
            kwargs['index'] = index
        self.book.create_sheet(**kwargs)


    # Set a new sheetname
    @is_writable
    def set_sheet_name(self, cur_name, new_name):
        self.book[cur_name].title = new_name


    # Set range values 
    @is_writable
    def set_range_values(self, sheetname, row, col, data_matrix):
        sheet = self.book[sheetname]
        for i in range(len(data_matrix)):
            for j in range(len(data_matrix[i])):
                sheet.cell(row=row+i, column=col+j).value = data_matrix[i][j]



def load_excel(file_path):
    if '/' in file_path:
        file_path = file_path.replace('/', os.sep)

    if not os.path.exists(file_path):
        raise IOError(file_path + ': No such file')

    return openpyxl.load_workbook(file_path)
